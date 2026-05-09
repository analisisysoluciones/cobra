from datetime import date
from decimal import Decimal

from django.http import HttpResponse
from django.shortcuts import render

from django.conf import settings
import os

from nomina.models import NominaDetalle, Empleado
from adm.models import Proyecto  # FK de nomina_empleado.proyecto


def construir_reporte_detallado_conceptos(inicio, fin, empleado_id=None, proyecto_id=None):
    """
    Construye estructura:
    [
        {
            "proyecto": "AMATITAN",
            "empleados": [
                {
                    "empleado": "Jesús Quiñones",
                    "percepciones": Decimal,
                    "deducciones": Decimal,
                    "neto": Decimal,
                    "conceptos_per": [
                        {"concepto": "...", "cantidad": ..., "monto_unitario": ..., "importe": ...},
                        ...
                    ],
                    "conceptos_ded": [...],
                },
                ...
            ],
            "total_percepciones": Decimal,
            "total_deducciones": Decimal,
            "total_neto": Decimal,
        },
        ...
    ]
    """

    qs = NominaDetalle.objects.filter(
        nomina_empleado__historial__periodo_nomina__periodo_inicio__range=[inicio, fin]
    ).select_related(
        "nomina_empleado__empleado",
        "nomina_empleado__proyecto",
        "nomina_empleado__historial__periodo_nomina",
    )

    if empleado_id:
        qs = qs.filter(nomina_empleado__empleado_id=empleado_id)

    if proyecto_id:
        qs = qs.filter(nomina_empleado__proyecto_id=proyecto_id)

    estructura = {}  # proyecto_nombre -> dict

    for d in qs:
        emp = d.nomina_empleado.empleado
        proy = d.nomina_empleado.proyecto

        proyecto_nombre = proy.nombre if proy else "SIN PROYECTO"

        # Proyecto
        p_data = estructura.setdefault(proyecto_nombre, {
            "proyecto": proyecto_nombre,
            "empleados": {},  # luego lo convertimos a lista
            "total_percepciones": Decimal("0.00"),
            "total_deducciones": Decimal("0.00"),
            "total_neto": Decimal("0.00"),
        })

        # Empleado dentro de proyecto
        e_data = p_data["empleados"].setdefault(emp.id, {
            "empleado": emp.nombre,
            "percepciones": Decimal("0.00"),
            "deducciones": Decimal("0.00"),
            "neto": Decimal("0.00"),
            "conceptos_per": [],
            "conceptos_ded": [],
        })

        concepto_info = {
            "concepto": d.concepto or "",
            "tipo": d.tipo,  # PERCEPCION / DEDUCCION
            "cantidad": d.cantidad,
            "monto_unitario": d.monto_unitario,
            "importe": d.subtotal,
        }

        if d.tipo == "PERCEPCION":
            e_data["percepciones"] += d.subtotal
            p_data["total_percepciones"] += d.subtotal
            e_data["conceptos_per"].append(concepto_info)
        else:
            e_data["deducciones"] += d.subtotal
            p_data["total_deducciones"] += d.subtotal
            e_data["conceptos_ded"].append(concepto_info)

    reporte = []

    # Ordenar por proyecto y dentro por empleado
    for proyecto_nombre in sorted(estructura.keys()):
        p_data = estructura[proyecto_nombre]

        empleados_dict = p_data["empleados"]
        empleados_list = []

        for _, e_data in sorted(empleados_dict.items(), key=lambda x: x[1]["empleado"]):
            # calcular neto empleado
            e_data["neto"] = e_data["percepciones"] - e_data["deducciones"]

            # ordenar conceptos: primero PERCEPCIONES, luego DEDUCCIONES
            # dentro de cada grupo, orden alfabético por nombre de concepto
            e_data["conceptos_per"] = sorted(
                e_data["conceptos_per"],
                key=lambda c: c["concepto"] or ""
            )
            e_data["conceptos_ded"] = sorted(
                e_data["conceptos_ded"],
                key=lambda c: c["concepto"] or ""
            )

            empleados_list.append(e_data)

        p_data["empleados"] = empleados_list
        p_data["total_neto"] = p_data["total_percepciones"] - p_data["total_deducciones"]

        reporte.append(p_data)

    return reporte



def reporte_nomina_detalle_view(request):
    empleados = Empleado.objects.all().order_by("nombre")
    proyectos = Proyecto.objects.all().order_by("nombre")

    empleado_id = request.GET.get("empleado") or None
    proyecto_id = request.GET.get("proyecto") or None
    fecha_desde = request.GET.get("fecha_desde")
    fecha_hasta = request.GET.get("fecha_hasta")

    if fecha_desde and fecha_hasta:
        inicio = date.fromisoformat(fecha_desde)
        fin = date.fromisoformat(fecha_hasta)
    else:
        inicio = date(2000, 1, 1)
        fin = date.today()

    reporte = construir_reporte_detallado_conceptos(
        inicio=inicio,
        fin=fin,
        empleado_id=empleado_id,
        proyecto_id=proyecto_id,
    )

    context = {
        "reporte": reporte,
        "empleados": empleados,
        "proyectos": proyectos,
        "inicio": inicio,
        "fin": fin,
        "request": request,
    }
    return render(request, "nomina/reporte_nomina_detalle.html", context)



from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib import colors


def reporte_nomina_detalle_pdf(request):
    empleado_id = request.GET.get("empleado") or None
    proyecto_id = request.GET.get("proyecto") or None
    fecha_desde = request.GET.get("fecha_desde")
    fecha_hasta = request.GET.get("fecha_hasta")

    if fecha_desde and fecha_hasta:
        inicio = date.fromisoformat(fecha_desde)
        fin = date.fromisoformat(fecha_hasta)
    else:
        inicio = date(2000, 1, 1)
        fin = date.today()

    reporte = construir_reporte_detallado_conceptos(
        inicio=inicio,
        fin=fin,
        empleado_id=empleado_id,
        proyecto_id=proyecto_id,
    )

    return generar_pdf_detalle_conceptos(reporte, inicio, fin)


def generar_pdf_detalle_conceptos(reporte, inicio, fin):
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="reporte_nomina_detalle_conceptos.pdf"'

    p = canvas.Canvas(response, pagesize=landscape(letter))
    width, height = landscape(letter)

    # Encabezado INEMO
    logo_path = os.path.join(settings.BASE_DIR, 'static', 'base', 'inemo.png')
    if os.path.exists(logo_path):
        p.drawImage(logo_path, 40, height - 90, width=120, height=60, preserveAspectRatio=True)

    p.setFont("Helvetica-Bold", 16)
    p.drawString(180, height - 50, "INEMO — Detalle de Nómina por Concepto")

    p.setFont("Helvetica", 10)
    p.drawString(180, height - 70, f"Periodo: {inicio} al {fin}")

    from datetime import datetime
    fecha_hora = datetime.now().strftime("%d/%m/%Y %H:%M")
    p.drawRightString(width - 40, height - 50, f"Generado: {fecha_hora}")

    y = height - 110

    def nueva_pagina():
        nonlocal y
        p.showPage()
        # Encabezado (sin repetir logo para simplificar, si quieres se repite)
        p.setFont("Helvetica-Bold", 16)
        p.drawString(40, height - 50, "INEMO — Detalle de Nómina por Concepto")
        p.setFont("Helvetica", 10)
        p.drawString(40, height - 70, f"Periodo: {inicio} al {fin}")
        y = height - 100

    # Recorremos proyectos
    for p_data in reporte:
        if y < 120:
            nueva_pagina()

        p.setFont("Helvetica-Bold", 12)
        p.drawString(40, y, f"PROYECTO: {p_data['proyecto']}")
        y -= 15

        # Recorremos empleados
        for e in p_data["empleados"]:
            if y < 120:
                nueva_pagina()
                p.setFont("Helvetica-Bold", 12)
                p.drawString(40, y, f"PROYECTO: {p_data['proyecto']}")
                y -= 15

            p.setFont("Helvetica-Bold", 10)
            p.drawString(60, y, f"EMPLEADO: {e['empleado']}")
            y -= 12

            p.setFont("Helvetica-Bold", 9)
            p.drawString(70, y, "Concepto")
            p.drawString(280, y, "Tipo")
            p.drawRightString(420, y, "Cantidad")
            p.drawRightString(520, y, "Monto Unitario")
            p.drawRightString(640, y, "Importe")
            y -= 10
            p.line(70, y, width - 40, y)
            y -= 8

            p.setFont("Helvetica", 9)

            # PERCEPCIONES
            if e["conceptos_per"]:
                p.setFont("Helvetica-Bold", 9)
                p.drawString(70, y, "PERCEPCIONES")
                y -= 12
                p.setFont("Helvetica", 9)

                for c in e["conceptos_per"]:
                    if y < 80:
                        nueva_pagina()
                    p.drawString(80, y, (c["concepto"] or "")[:40])
                    p.drawString(280, y, "PERCEPCIÓN")
                    p.drawRightString(420, y, f"{c['cantidad']:.2f}")
                    p.drawRightString(520, y, f"{c['monto_unitario']:.2f}")
                    p.drawRightString(640, y, f"{c['importe']:.2f}")
                    y -= 12

            # DEDUCCIONES
            if e["conceptos_ded"]:
                if y < 80:
                    nueva_pagina()
                p.setFont("Helvetica-Bold", 9)
                p.drawString(70, y, "DEDUCCIONES")
                y -= 12
                p.setFont("Helvetica", 9)

                for c in e["conceptos_ded"]:
                    if y < 80:
                        nueva_pagina()
                    p.drawString(80, y, (c["concepto"] or "")[:40])
                    p.drawString(280, y, "DEDUCCIÓN")
                    p.drawRightString(420, y, f"{c['cantidad']:.2f}")
                    p.drawRightString(520, y, f"{c['monto_unitario']:.2f}")
                    p.drawRightString(640, y, f"{c['importe']:.2f}")
                    y -= 12

            # Subtotal empleado
            if y < 80:
                nueva_pagina()
            p.setFont("Helvetica-Bold", 9)
            p.drawRightString(520, y, "Subtotal empleado (P/D):")
            p.drawRightString(640, y, f"{e['percepciones']:.2f} / {e['deducciones']:.2f}")
            y -= 12
            p.drawRightString(640, y, f"Neto: {e['neto']:.2f}")
            y -= 18

        # Subtotal proyecto
        if y < 80:
            nueva_pagina()
        p.setFont("Helvetica-Bold", 10)
        p.drawRightString(520, y, f"Subtotal proyecto {p_data['proyecto']}:")
        y -= 12
        p.drawRightString(640, y, f"Percepciones: {p_data['total_percepciones']:.2f}")
        y -= 12
        p.drawRightString(640, y, f"Deducciones: {p_data['total_deducciones']:.2f}")
        y -= 12
        p.drawRightString(640, y, f"Neto: {p_data['total_neto']:.2f}")
        y -= 24

    p.showPage()
    p.save()
    return response



from openpyxl import Workbook
from openpyxl.styles import Font, Alignment


def reporte_nomina_detalle_excel(request):

    empleado_id = request.GET.get("empleado") or None
    proyecto_id = request.GET.get("proyecto") or None
    fecha_desde = request.GET.get("fecha_desde")
    fecha_hasta = request.GET.get("fecha_hasta")

    if fecha_desde and fecha_hasta:
        inicio = date.fromisoformat(fecha_desde)
        fin = date.fromisoformat(fecha_hasta)
    else:
        inicio = date(2000, 1, 1)
        fin = date.today()

    reporte = construir_reporte_detallado_conceptos(
        inicio=inicio,
        fin=fin,
        empleado_id=empleado_id,
        proyecto_id=proyecto_id,
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Detalle Conceptos"

    # Encabezado grande INEMO
    ws["A1"] = "INEMO - Reporte Detallado por Concepto"
    ws["A1"].font = Font(bold=True, size=16)
    ws.merge_cells("A1:F1")

    ws["A2"] = f"Periodo: {inicio} al {fin}"
    ws["A2"].font = Font(italic=True)
    ws.merge_cells("A2:F2")

    row = 4

    for p in reporte:
        ws[f"A{row}"] = f"PROYECTO: {p['proyecto']}"
        ws[f"A{row}"].font = Font(bold=True, size=12, color="1f4e79")
        row += 2

        for e in p["empleados"]:
            ws[f"A{row}"] = f"EMPLEADO: {e['empleado']}"
            ws[f"A{row}"].font = Font(bold=True, size=11)
            row += 1

            # Encabezados tabla
            headers = ["Concepto", "Tipo", "Cantidad", "Monto Unitario", "Importe"]
            ws.append(headers)
            for col in range(1, 6):
                ws.cell(row=row, column=col).font = Font(bold=True)
            row += 1

            # Percepciones
            if e["conceptos_per"]:
                ws[f"A{row}"] = "PERCEPCIONES"
                ws[f"A{row}"].font = Font(bold=True)
                row += 1

                for c in e["conceptos_per"]:
                    ws.append([
                        c["concepto"],
                        "PERCEPCION",
                        c["cantidad"],
                        float(c["monto_unitario"]),
                        float(c["importe"]),
                    ])
                    row += 1

            # Deducciones
            if e["conceptos_ded"]:
                ws[f"A{row}"] = "DEDUCCIONES"
                ws[f"A{row}"].font = Font(bold=True)
                row += 1

                for c in e["conceptos_ded"]:
                    ws.append([
                        c["concepto"],
                        "DEDUCCION",
                        c["cantidad"],
                        float(c["monto_unitario"]),
                        float(c["importe"]),
                    ])
                    row += 1

            # subtotal empleado
            ws[f"A{row}"] = "Subtotal empleado"
            ws[f"E{row}"] = float(e["neto"])
            ws[f"A{row}"].font = Font(bold=True)
            ws[f"E{row}"].font = Font(bold=True)
            row += 2

        # subtotal proyecto
        ws[f"A{row}"] = f"Subtotal proyecto {p['proyecto']}"
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"E{row}"] = float(p["total_neto"])
        ws[f"E{row}"].font = Font(bold=True)
        row += 3

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = "attachment; filename=reporte_detallado_conceptos.xlsx"

    wb.save(response)
    return response
