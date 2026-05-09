from django.db.models import Sum, F, DecimalField, ExpressionWrapper
from datetime import date, timedelta, datetime
from django.shortcuts import render
# nomina/utils_pdf.py
from reportlab.lib import colors
from reportlab.platypus import Table, TableStyle
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter, landscape
from django.http import HttpResponse
# nomina/utils_excel.py
from openpyxl import Workbook
from openpyxl.styles import Font
from decimal import Decimal
import os
from django.conf import settings

from nomina.models import (
    Empleado, PeriodosNomina, NominaDetalle, NominaEmpleado,
    HorasExtras, CompensacionVariable, AsistenciaDia, RegistroDestajo
)

from adm.models import Proyecto





def reporte_nomina_view(request):
    # --- Combos ---
    empleados = Empleado.objects.all().order_by("nombre")
    proyectos = Proyecto.objects.all().order_by("nombre")   # Si es Obra, cambia a Obra

    # --- Filtros ---
    empleado_id = request.GET.get("empleado") or None
    proyecto_id = request.GET.get("proyecto") or None

    fecha_desde = request.GET.get("fecha_desde")
    fecha_hasta = request.GET.get("fecha_hasta")

    # Si NO elige rango → no filtramos por fecha todavía
    if fecha_desde and fecha_hasta:
        inicio = date.fromisoformat(fecha_desde)
        fin = date.fromisoformat(fecha_hasta)
        etiqueta_periodo = f"{inicio} a {fin}"
    else:
        # Mostrar TODO por defecto (si lo prefieres puedo poner últimos 7 días)
        inicio = date(2000, 1, 1)
        fin = date.today()
        etiqueta_periodo = "Todos los periodos"

    # Construimos el resumen unificado
    resumen = construir_resumen_nomina(
        inicio=inicio,
        fin=fin,
        empleado_id=empleado_id,
        proyecto_id=proyecto_id,
    )

    context = {
        "resumen": resumen,
        "empleados": empleados,
        "proyectos": proyectos,
        "inicio": inicio,
        "fin": fin,
        "etiqueta_periodo": etiqueta_periodo,
    }
    return render(request, "nomina/reporte_nomina.html", context)




def generar_reporte_nomina(filtros):
    qs = NominaDetalle.objects.select_related(
        'nomina', 'empleado', 'proyecto', 'concepto', 'nomina__periodo_nomina'
    )

    # Filtros dinámicos
    if filtros.get('fecha_desde') and filtros.get('fecha_hasta'):
        qs = qs.filter(
            nomina__periodo_nomina__periodo_inicio__gte=filtros['fecha_desde'],
            nomina__periodo_nomina__periodo_final__lte=filtros['fecha_hasta']
        )

    if filtros.get('anio'):
        qs = qs.filter(nomina__periodo_nomina__anio=filtros['anio'])
    if filtros.get('mes'):
        qs = qs.filter(nomina__periodo_nomina__periodo_inicio__month=filtros['mes'])
    if filtros.get('empleado'):
        qs = qs.filter(empleado_id=filtros['empleado'])
    if filtros.get('proyecto'):
        qs = qs.filter(proyecto_id=filtros['proyecto'])
    if filtros.get('tipo'):
        qs = qs.filter(concepto__tipo=filtros['tipo'])
    if filtros.get('estatus'):
        qs = qs.filter(nomina__estatus=filtros['estatus'])

    # Agrupación para reporte
    resumen = (
        qs.values(
            'nomina__periodo_nomina__anio',
            'nomina__periodo_nomina__semana',
            'empleado__nombre',
            'proyecto__nombre',
            'concepto__descripcion',
            'concepto__tipo',
        )
        .annotate(total=Sum('importe'))
        .order_by('nomina__periodo_nomina__anio', 'empleado__nombre')
    )

    return resumen



def resumen_totales(qs):
    percepciones = qs.filter(concepto__tipo='PERCEPCION').aggregate(total=Sum('importe'))['total'] or 0
    deducciones = qs.filter(concepto__tipo='DEDUCCION').aggregate(total=Sum('importe'))['total'] or 0
    neto = percepciones - abs(deducciones)
    return {
        'percepciones': percepciones,
        'deducciones': deducciones,
        'neto': neto,
    }



def reporte_nomina_pdf(request):
    empleado_id = request.GET.get("empleado") or None
    proyecto_id = request.GET.get("proyecto") or None

    fecha_desde = request.GET.get("fecha_desde")
    fecha_hasta = request.GET.get("fecha_hasta")

    # Rango opcional
    if fecha_desde and fecha_hasta:
        inicio = date.fromisoformat(fecha_desde)
        fin = date.fromisoformat(fecha_hasta)
    else:
        inicio = date(2000, 1, 1)
        fin = date.today()

    # Obtener resumen
    resumen = construir_resumen_nomina(
        inicio=inicio,
        fin=fin,
        empleado_id=empleado_id,
        proyecto_id=proyecto_id,
    )

    return generar_pdf_resumen(resumen, inicio, fin)

def reporte_nomina_excel(request):
    empleado_id = request.GET.get("empleado") or None
    proyecto_id = request.GET.get("proyecto") or None

    inicio, fin, etiqueta_periodo = resolver_rango_periodo(request)

    resumen = construir_resumen_nomina(
        inicio=inicio,
        fin=fin,
        empleado_id=empleado_id,
        proyecto_id=proyecto_id,
    )

    return generar_excel_resumen(resumen, inicio, fin, etiqueta_periodo)



def generar_pdf(resultados, filtros):
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="reporte_nomina.pdf"'

    p = canvas.Canvas(response, pagesize=landscape(letter))
    width, height = landscape(letter)

    y = height - 50

    p.setFont("Helvetica-Bold", 16)
    p.drawString(50, y, "Reporte de Nómina")
    y -= 40

    p.setFont("Helvetica", 10)

    # Filtros usados
    p.drawString(50, y, f"Filtros aplicados: {dict(filtros)}")
    y -= 30

    # Encabezados
    p.setFont("Helvetica-Bold", 10)
    p.drawString(50, y, "Año")
    p.drawString(100, y, "Semana")
    p.drawString(160, y, "Empleado")
    p.drawString(360, y, "Proyecto")
    p.drawString(520, y, "Concepto")
    p.drawString(650, y, "Tipo")
    p.drawString(720, y, "Importe")
    y -= 20

    p.setFont("Helvetica", 10)

    for r in resultados:
        if y < 50:
            p.showPage()
            y = height - 50

        p.drawString(50, y, str(r['nomina__periodo_nomina__anio']))
        p.drawString(100, y, str(r['nomina__periodo_nomina__semana']))
        p.drawString(160, y, str(r['empleado__nombre']))
        p.drawString(360, y, r.get('proyecto__nombre', '') or '')
        p.drawString(520, y, r['concepto__descripcion'])
        p.drawString(650, y, r['concepto__tipo'])
        p.drawString(720, y, f"{r['total']:.2f}")
        y -= 20

    p.showPage()
    p.save()

    return response




def generar_excel(resultados, filtros):
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte Nómina"

    # Encabezados
    encabezados = [
        "Año", "Semana", "Empleado", "Proyecto",
        "Concepto", "Tipo", "Importe"
    ]
    ws.append(encabezados)

    # Negritas
    for cell in ws[1]:
        cell.font = Font(bold=True)

    # Renglones
    for r in resultados:
        ws.append([
            r['nomina__periodo_nomina__anio'],
            r['nomina__periodo_nomina__semana'],
            r['empleado__nombre'],
            r.get('proyecto__nombre', ''),
            r['concepto__descripcion'],
            r['concepto__tipo'],
            r['total']
        ])

    # Respuesta Http
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = 'attachment; filename="reporte_nomina.xlsx"'
    wb.save(response)
    return response








def resolver_rango_periodo(request):
    """
    Devuelve (inicio, fin, etiqueta_periodo) en base a los filtros:
    - modo: 'rango' | 'semanal' | 'mensual' | 'anual'
    - fecha_desde, fecha_hasta
    - anio, mes, semana
    """
    modo = request.GET.get("modo", "rango")
    fecha_desde = request.GET.get("fecha_desde")
    fecha_hasta = request.GET.get("fecha_hasta")
    anio = request.GET.get("anio")
    mes = request.GET.get("mes")
    semana = request.GET.get("semana")

    hoy = date.today()

    # -------------------
    # RANGO MANUAL
    # -------------------
    if modo == "rango":
        if fecha_desde:
            inicio = date.fromisoformat(fecha_desde)
        else:
            inicio = hoy - timedelta(days=7)

        if fecha_hasta:
            fin = date.fromisoformat(fecha_hasta)
        else:
            fin = hoy

        etiqueta = f"Rango manual: {inicio} al {fin}"
        return inicio, fin, etiqueta

    # -------------------
    # ANUAL
    # -------------------
    if modo == "anual" and anio:
        anio_int = int(anio)
        inicio = date(anio_int, 1, 1)
        fin = date(anio_int, 12, 31)
        etiqueta = f"Año {anio_int}"
        return inicio, fin, etiqueta

    # -------------------
    # MENSUAL
    # -------------------
    if modo == "mensual" and anio and mes:
        anio_int = int(anio)
        mes_int = int(mes)

        inicio = date(anio_int, mes_int, 1)
        # calcular último día del mes
        if mes_int == 12:
            siguiente = date(anio_int + 1, 1, 1)
        else:
            siguiente = date(anio_int, mes_int + 1, 1)
        fin = siguiente - timedelta(days=1)
        etiqueta = f"Mes {mes_int:02d}/{anio_int}"
        return inicio, fin, etiqueta

    # -------------------
    # SEMANAL (usando PeriodosNomina)
    # -------------------
    if modo == "semanal" and anio and semana:
        try:
            per = PeriodosNomina.objects.get(anio=int(anio), semana=int(semana))
            inicio = per.periodo_inicio
            fin = per.periodo_final
            etiqueta = f"Semana {per.semana} ({inicio} al {fin})"
            return inicio, fin, etiqueta
        except PeriodosNomina.DoesNotExist:
            # si no existe, cae al default
            pass

    # -------------------
    # DEFAULT (últimos 7 días)
    # -------------------
    fin = hoy
    inicio = hoy - timedelta(days=7)
    etiqueta = f"Últimos 7 días ({inicio} al {fin})"
    return inicio, fin, etiqueta



def construir_resumen_nomina(inicio, fin, empleado_id=None, proyecto_id=None):
    """
    Regresa una lista de diccionarios, uno por (empleado, proyecto):
    {
        'empleado': str,
        'proyecto': str,
        'percepciones': Decimal,
        'deducciones': Decimal,
        'destajos': Decimal,
        'horas_extra': Decimal,
        'compensaciones': Decimal,
        'faltas': int,
        'total_neto': Decimal,
    }
    """
    resumen = {}

    def get_or_create_row(emp, proy_texto):
        key = (emp.id if emp else None, proy_texto)
        if key not in resumen:
            resumen[key] = {
                "empleado": emp.nombre if emp else "N/D",
                "proyecto": proy_texto,
                "percepciones": Decimal("0.00"),
                "deducciones": Decimal("0.00"),
                "destajos": Decimal("0.00"),
                "horas_extra": Decimal("0.00"),
                "compensaciones": Decimal("0.00"),
                "faltas": 0,
                "total_neto": Decimal("0.00"),
            }
        return resumen[key]

    # -------------------------------------------------
    # 1) PERCEPCIONES / DEDUCCIONES DE NOMINADETALLE
    # -------------------------------------------------
    detalles = NominaDetalle.objects.filter(
        nomina_empleado__historial__periodo_nomina__periodo_inicio__range=[inicio, fin]
    )

    if empleado_id:
        detalles = detalles.filter(nomina_empleado__empleado_id=empleado_id)
    if proyecto_id:
        detalles = detalles.filter(nomina_empleado__proyecto_id=proyecto_id)

    for d in detalles.select_related(
        "nomina_empleado__empleado",
        "nomina_empleado__proyecto",
        "nomina_empleado__historial__periodo_nomina",
    ):
        emp = d.nomina_empleado.empleado
        proy = d.nomina_empleado.proyecto
        proy_texto = proy.nombre if proy else "SIN PROYECTO"

        row = get_or_create_row(emp, proy_texto)

        if d.tipo == "PERCEPCION":
            row["percepciones"] += d.subtotal
        else:
            row["deducciones"] += d.subtotal

    # -------------------------------------------------
    # 2) HORAS EXTRA (HorasExtras)
    # -------------------------------------------------
    horas_extra = HorasExtras.objects.filter(
        periodo__periodo_inicio__range=[inicio, fin]
    )
    if empleado_id:
        horas_extra = horas_extra.filter(empleado_id=empleado_id)
    if proyecto_id:
        horas_extra = horas_extra.filter(proyecto_id=proyecto_id)

    for h in horas_extra.select_related("empleado", "proyecto", "periodo"):
        emp = h.empleado
        proy = h.proyecto
        proy_texto = proy.nombre if proy else "SIN PROYECTO"

        row = get_or_create_row(emp, proy_texto)
        row["horas_extra"] += h.total_pago

    # -------------------------------------------------
    # 3) COMPENSACIONES VARIABLES
    # -------------------------------------------------
    comps = CompensacionVariable.objects.filter(
        periodo__periodo_inicio__range=[inicio, fin]
    )
    if empleado_id:
        comps = comps.filter(empleado_id=empleado_id)
    if proyecto_id:
        comps = comps.filter(proyecto_id=proyecto_id)

    for c in comps.select_related("empleado", "proyecto", "periodo"):
        emp = c.empleado
        proy = c.proyecto
        proy_texto = proy.nombre if proy else "SIN PROYECTO"

        row = get_or_create_row(emp, proy_texto)
        row["compensaciones"] += c.monto

    # -------------------------------------------------
    # 4) DESTAJO (RegistroDestajo)
    # -------------------------------------------------
    destajos = RegistroDestajo.objects.filter(
        semana__periodo_inicio__range=[inicio, fin]
    )
    if empleado_id:
        destajos = destajos.filter(empleado_id=empleado_id)
    if proyecto_id:
        destajos = destajos.filter(obra_id=proyecto_id)

    for d in destajos.select_related("empleado", "obra", "semana"):
        emp = d.empleado  # puede ser null
        obra = d.obra
        proy_texto = obra.nombre

        row = get_or_create_row(emp, proy_texto)
        row["destajos"] += d.total

    # -------------------------------------------------
    # 5) FALTAS (AsistenciaDia.laboro == 0)
    # -------------------------------------------------
    asistencias = AsistenciaDia.objects.filter(
        semana__periodo_inicio__range=[inicio, fin]
    )
    if empleado_id:
        asistencias = asistencias.filter(empleado_id=empleado_id)
    if proyecto_id:
        asistencias = asistencias.filter(obra_id=proyecto_id)

    for a in asistencias.select_related("empleado", "obra", "semana"):
        try:
            laboro = Decimal(a.laboro)
        except (TypeError, ValueError):
            laboro = Decimal("0.00")

        if laboro == 0:
            emp = a.empleado
            obra = a.obra
            proy_texto = obra.nombre
            row = get_or_create_row(emp, proy_texto)
            row["faltas"] += 1

    # -------------------------------------------------
    # 6) TOTAL NETO
    # -------------------------------------------------
    for row in resumen.values():
        row["total_neto"] = (
            row["percepciones"]
            + row["destajos"]
            + row["horas_extra"]
            + row["compensaciones"]
            - row["deducciones"]
        )

    # Lo devolvemos ordenado por proyecto y empleado
    resumen_ordenado = sorted(
        resumen.values(),
        key=lambda x: (x["proyecto"], x["empleado"])
    )
    return resumen_ordenado



def generar_pdf_resumen(resumen, inicio, fin):

    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import letter, landscape
    from reportlab.lib import colors
    from reportlab.platypus import Table, TableStyle
    from reportlab.lib.units import cm
    from django.conf import settings
    import os
    from datetime import datetime

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="reporte_nomina_resumen.pdf"'

    p = canvas.Canvas(response, pagesize=landscape(letter))
    width, height = landscape(letter)

    # -----------------------------------------------
    # ENCABEZADO CORPORATIVO INEMO
    # -----------------------------------------------
    logo_path = os.path.join(settings.BASE_DIR, 'static', 'base', 'inemo.png')

    if os.path.exists(logo_path):
        p.drawImage(logo_path, 40, height - 90, width=120, height=60, preserveAspectRatio=True)

    p.setFont("Helvetica-Bold", 18)
    p.drawString(180, height - 50, "INEMO — Reporte de Nómina Resumida")

    p.setFont("Helvetica", 10)
    p.drawString(180, height - 70, f"Periodo: {inicio} al {fin}")

    fecha_hora = datetime.now().strftime("%d/%m/%Y %H:%M")
    p.drawRightString(width - 40, height - 50, f"Generado: {fecha_hora}")

    y = height - 120  # posición inicial para el contenido

    # -----------------------------------------------
    # AGRUPAR POR PROYECTO
    # -----------------------------------------------
    proyectos = {}
    for r in resumen:
        proyectos.setdefault(r["proyecto"], []).append(r)

    # -----------------------------------------------
    # RECORRER PROYECTOS
    # -----------------------------------------------
    for proyecto, elementos in proyectos.items():

        # ENCABEZADO DE PROYECTO
        p.setFont("Helvetica-Bold", 14)
        p.drawString(40, y, f"PROYECTO: {proyecto}")
        y -= 20

        # Encabezado de la tabla (sin “Proyecto”)
        data = [[
            "Empleado", "Percepciones", "Deducciones",
            "Destajos", "Horas Extra", "Compensaciones", "Faltas", "Total Neto"
        ]]

        # Filas del proyecto
        for r in elementos:
            data.append([
                r["empleado"],
                f"${r['percepciones']:.2f}",
                f"${r['deducciones']:.2f}",
                f"${r['destajos']:.2f}",
                f"${r['horas_extra']:.2f}",
                f"${r['compensaciones']:.2f}",
                r["faltas"],
                f"${r['total_neto']:.2f}",
            ])

        # SUBTOTALES DEL PROYECTO
        total_per = sum([row["percepciones"] for row in elementos])
        total_ded = sum([row["deducciones"] for row in elementos])
        total_des = sum([row["destajos"] for row in elementos])
        total_he  = sum([row["horas_extra"] for row in elementos])
        total_cv  = sum([row["compensaciones"] for row in elementos])
        total_fal = sum([row["faltas"] for row in elementos])
        total_net = sum([row["total_neto"] for row in elementos])

        data.append([
            "SUBTOTAL",
            f"${total_per:.2f}",
            f"${total_ded:.2f}",
            f"${total_des:.2f}",
            f"${total_he:.2f}",
            f"${total_cv:.2f}",
            total_fal,
            f"${total_net:.2f}",
        ])

        # Estilo y ancho de tabla
        table = Table(data, colWidths=[150, 70, 70, 70, 70, 90, 40, 90])

        style = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#1f4e79")),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),

            ('ALIGN', (1, 1), (-1, -1), 'RIGHT'),
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),

            ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),

            ('GRID', (0, 0), (-1, -1), 0.3, colors.grey),
            ('LEFTPADDING', (0, 0), (-1, -1), 4),
            ('RIGHTPADDING', (0, 0), (-1, -1), 4),
        ])

        table.setStyle(style)

        # Si se llena la página, nueva página
        if y < 200:
            p.showPage()
            y = height - 100

        table.wrapOn(p, width, height)
        table.drawOn(p, 40, y - (len(data) * 18))

        y -= (len(data) * 18 + 40)

    # FIN DE REPORTE
    p.showPage()
    p.save()

    return response




def generar_excel_resumen(resumen, inicio, fin, etiqueta_periodo):
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte Nómina"

    ws["A1"] = "Reporte de Nómina - Resumen"
    ws["A2"] = f"Periodo: {inicio} al {fin} ({etiqueta_periodo})"
    ws["A1"].font = Font(bold=True)
    ws["A2"].font = Font(bold=True)

    # Encabezados
    encabezados = [
        "Empleado", "Proyecto",
        "Percepciones", "Deducciones",
        "Destajos", "Horas Extra",
        "Compensaciones", "Faltas",
        "Total Neto",
    ]
    ws.append([])
    ws.append(encabezados)

    for cell in ws[4]:
        cell.font = Font(bold=True)

    for r in resumen:
        ws.append([
            r["empleado"],
            r["proyecto"],
            float(r["percepciones"]),
            float(r["deducciones"]),
            float(r["destajos"]),
            float(r["horas_extra"]),
            float(r["compensaciones"]),
            r["faltas"],
            float(r["total_neto"]),
        ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = 'attachment; filename=\"reporte_nomina_resumen.xlsx\"'
    wb.save(response)
    return response
